from __future__ import annotations

import argparse
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from titan007_extract_euro_odds import (
    build_plugin_company_rows,
    fetch_oddslist_js,
    parse_oddslist_js,
)


ROOT_DIR = Path(__file__).resolve().parents[1]
PLUGIN_DIR = ROOT_DIR / "runtime" / "football-odds-predictor"
if str(PLUGIN_DIR) not in sys.path:
    sys.path.insert(0, str(PLUGIN_DIR))

from predictor_py import (  # noqa: E402
    FIXED_COMPANIES,
    compute_report_prediction,
    get_betting_decision_table,
    get_live_mode_selection_table,
)


OUTPUT_DIR = ROOT_DIR / "output" / "spreadsheet"
TIMEZONE_BEIJING = ZoneInfo("Asia/Shanghai")
WINDOW_START_BJT = datetime(2026, 3, 29, 18, 0, tzinfo=TIMEZONE_BEIJING)
WINDOW_END_BJT = datetime(2026, 4, 3, 0, 0, tzinfo=TIMEZONE_BEIJING)
CURRENT_TIME_BJT = datetime.now(TIMEZONE_BEIJING)
FUTURE_SCHEDULE_DATES = ["20260329", "20260330", "20260331", "20260401", "20260402"]
FUTURE_SCHEDULE_URL_TEMPLATE = "https://bf.titan007.com/football/Next_{date}.htm"
INITIAL_MAX_WORKERS = 8
RETRY_MAX_WORKERS = 3
ALLOWED_CONFIDENCES = {"高", "中"}
LARGE_WINDOW_MATCH_THRESHOLD = 500
LARGE_WINDOW_MAX_WORKERS = 24
LARGE_WINDOW_RETRY_MAX_WORKERS = 8
VALID_CONFIDENCES = {"高", "中", "谨慎"}


@dataclass
class FutureMatch:
    schedule_id: str
    league: str
    kickoff_bjt: datetime
    home_team: str
    away_team: str
    europe_odds_url: str
    source_page_url: str


@dataclass
class MatchReportRow:
    league: str
    kickoff_bjt: datetime
    kickoff_text: str
    home_team: str
    away_team: str
    recommendation: str
    structure_label: str
    confidence: str
    consensus: float
    top_gap: float
    phase_status: str
    effective_mode: str
    final_action: str
    final_prediction: str
    decision_basis: str
    mode_selection_basis: str
    mode_history_accuracy: str
    history_sample_size: str
    history_accuracy_summary: str
    source_match_id: str
    source_page_url: str
    source_js_url: str
    opening_odds_json: str
    closing_odds_json: str
    explanation: str


def parse_datetime_bjt(value: str) -> datetime:
    return datetime.strptime(value.strip(), "%Y-%m-%d %H:%M").replace(
        tzinfo=TIMEZONE_BEIJING
    )


def build_future_schedule_dates(start_bjt: datetime, end_bjt: datetime) -> list[str]:
    dates: list[str] = []
    current = start_bjt.date()
    end_date = end_bjt.date()
    while current <= end_date:
        dates.append(current.strftime("%Y%m%d"))
        current = current.fromordinal(current.toordinal() + 1)
    return dates


def parse_kickoff_bjt(raw_value: str) -> datetime | None:
    parts = [segment.strip() for segment in raw_value.split(",")]
    if len(parts) != 6:
        return None
    try:
        year = int(parts[0])
        month_zero_based = int(parts[1])
        day = int(parts[2])
        hour = int(parts[3])
        minute = int(parts[4])
        second = int(parts[5])
    except ValueError:
        return None
    return datetime(
        year,
        month_zero_based + 1,
        day,
        hour,
        minute,
        second,
        tzinfo=TIMEZONE_BEIJING,
    )


def fetch_future_schedule_html(date_text: str) -> str:
    url = FUTURE_SCHEDULE_URL_TEMPLATE.format(date=date_text)
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
    response.raise_for_status()
    return response.content.decode("gb18030", errors="ignore")


def clean_team_name(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def extract_team_name(cell_text: str) -> str:
    text = clean_team_name(cell_text)
    text = re.sub(r"^\[[^\]]+\]\s*", "", text)
    text = re.sub(r"\s*\[[^\]]+\]$", "", text)
    return text.strip()


def parse_future_schedule_matches() -> list[FutureMatch]:
    matches: list[FutureMatch] = []
    seen_ids: set[str] = set()

    for date_text in FUTURE_SCHEDULE_DATES:
        page_url = FUTURE_SCHEDULE_URL_TEMPLATE.format(date=date_text)
        html = fetch_future_schedule_html(date_text)
        soup = BeautifulSoup(html, "html.parser")

        for tr in soup.find_all("tr", sid=True):
            schedule_id = clean_team_name(tr.get("sid", ""))
            if not schedule_id or schedule_id in seen_ids:
                continue

            cells = tr.find_all("td")
            if len(cells) < 10:
                continue

            league = clean_team_name(" ".join(cells[0].stripped_strings))
            time_text = clean_team_name(" ".join(cells[1].stripped_strings))
            home_team = extract_team_name(" ".join(cells[3].stripped_strings))
            away_team = extract_team_name(" ".join(cells[5].stripped_strings))

            try:
                month_day, clock_text = time_text.split()
                month, day = [int(part) for part in month_day.split("-")]
                hour, minute = [int(part) for part in clock_text.split(":")]
                kickoff_bjt = datetime(
                    2026,
                    month,
                    day,
                    hour,
                    minute,
                    tzinfo=TIMEZONE_BEIJING,
                )
            except ValueError:
                continue

            if not (WINDOW_START_BJT <= kickoff_bjt <= WINDOW_END_BJT):
                continue

            seen_ids.add(schedule_id)
            matches.append(
                FutureMatch(
                    schedule_id=schedule_id,
                    league=league,
                    kickoff_bjt=kickoff_bjt,
                    home_team=home_team,
                    away_team=away_team,
                    europe_odds_url=f"https://1x2.titan007.com/oddslist/{schedule_id}.htm",
                    source_page_url=page_url,
                )
            )

    matches.sort(key=lambda item: (item.kickoff_bjt, item.league, item.home_team, item.away_team))
    return matches


def format_kickoff_bjt(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M")


def format_confidence_scope() -> str:
    ordered = [label for label in ("高", "中", "谨慎") if label in ALLOWED_CONFIDENCES]
    return "或".join(ordered) if ordered else "未指定"


def resolve_worker_counts(match_count: int) -> tuple[int, int]:
    if match_count >= LARGE_WINDOW_MATCH_THRESHOLD:
        return LARGE_WINDOW_MAX_WORKERS, LARGE_WINDOW_RETRY_MAX_WORKERS
    return INITIAL_MAX_WORKERS, RETRY_MAX_WORKERS


def normalize_sheet_name(name: str) -> str:
    cleaned = name.replace("/", "／").replace("\\", "＼").replace("*", "＊")
    cleaned = cleaned.replace(":", "：").replace("?", "？").replace("[", "［").replace("]", "］")
    return cleaned[:31] or "未命名联赛"


def workbook_header_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor="D9EAF7")


def section_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor="EDEDED")


def iter_report_rows() -> tuple[list[MatchReportRow], list[dict[str, str]]]:
    rows: list[MatchReportRow] = []
    audit_rows: list[dict[str, str]] = []
    future_matches = parse_future_schedule_matches()
    initial_max_workers, retry_max_workers = resolve_worker_counts(len(future_matches))
    report_rows, first_pass_audits = process_future_matches(
        future_matches,
        max_workers=initial_max_workers,
    )
    rows.extend(report_rows)
    audit_rows.extend(first_pass_audits)

    failed_match_ids = {
        item["schedule_id"]
        for item in first_pass_audits
        if item["status"] == "fetch_failed"
    }
    if failed_match_ids:
        retry_matches = [
            item for item in future_matches if item.schedule_id in failed_match_ids
        ]
        retry_rows, retry_audits = process_future_matches(
            retry_matches,
            max_workers=retry_max_workers,
        )
        rows.extend(retry_rows)

        filtered_first_pass_audits: list[dict[str, str]] = []
        for item in audit_rows:
            if item["schedule_id"] in failed_match_ids:
                continue
            filtered_first_pass_audits.append(item)
        audit_rows = filtered_first_pass_audits + retry_audits

    deduped_rows: dict[str, MatchReportRow] = {}
    for row in rows:
        deduped_rows[row.source_match_id] = row

    rows = sorted(
        deduped_rows.values(),
        key=lambda item: (item.league, item.kickoff_bjt, item.home_team, item.away_team),
    )
    return rows, audit_rows


def process_future_matches(
    matches: list[FutureMatch],
    max_workers: int,
) -> tuple[list[MatchReportRow], list[dict[str, str]]]:
    rows: list[MatchReportRow] = []
    audit_rows: list[dict[str, str]] = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for report_row, match_audit_rows in executor.map(process_future_match, matches):
            if report_row is not None:
                rows.append(report_row)
            audit_rows.extend(match_audit_rows)
    return rows, audit_rows


def process_future_match(match: FutureMatch) -> tuple[MatchReportRow | None, list[dict[str, str]]]:
    audit_rows: list[dict[str, str]] = []

    try:
        oddslist_js_url, oddslist_js = fetch_oddslist_js(match.schedule_id)
        oddslist_records = parse_oddslist_js(oddslist_js)
        plugin_rows, candidate_rows, missing = build_plugin_company_rows(
            oddslist_records,
            [],
        )
    except Exception as error:  # noqa: BLE001
        return None, [
            {
                "schedule_id": match.schedule_id,
                "status": "fetch_failed",
                "detail": f"{type(error).__name__}: {error}",
            }
        ]

    if missing:
        return None, [
            {
                "schedule_id": match.schedule_id,
                "status": "missing_companies",
                "detail": ",".join(missing),
            }
        ]

    opening_rows = [
        {
            "home": item["openingOdds"]["home"],
            "draw": item["openingOdds"]["draw"],
            "away": item["openingOdds"]["away"],
        }
        for item in plugin_rows
    ]
    closing_rows = []
    closing_available = True
    for item in plugin_rows:
        closing_odds = item.get("closingOdds")
        if not closing_odds:
            closing_available = False
            break
        closing_rows.append(
            {
                "home": closing_odds["home"],
                "draw": closing_odds["draw"],
                "away": closing_odds["away"],
            }
        )
    if not closing_available:
        closing_rows = None

    try:
        prediction = compute_report_prediction(
            opening_rows,
            closing_rows,
            match_started=CURRENT_TIME_BJT >= match.kickoff_bjt,
        )
    except Exception as error:  # noqa: BLE001
        return None, [
            {
                "schedule_id": match.schedule_id,
                "status": "prediction_failed",
                "detail": f"{type(error).__name__}: {error}",
            }
        ]

    opening_prediction = prediction["openingPrediction"]

    if opening_prediction["confidence"] not in ALLOWED_CONFIDENCES:
        return None, [
            {
                "schedule_id": match.schedule_id,
                "status": "confidence_not_high",
                "detail": opening_prediction["confidenceProfile"]["label"],
            }
        ]

    if candidate_rows:
        audit_rows.append(
            {
                "schedule_id": match.schedule_id,
                "status": "candidate_rows_present",
                "detail": json.dumps(candidate_rows, ensure_ascii=False),
            }
        )

    return (
        MatchReportRow(
            league=match.league,
            kickoff_bjt=match.kickoff_bjt,
            kickoff_text=format_kickoff_bjt(match.kickoff_bjt),
            home_team=match.home_team,
            away_team=match.away_team,
            recommendation=opening_prediction["recommendation"],
            structure_label=opening_prediction["confidenceProfile"]["label"],
            confidence=opening_prediction["confidence"],
            consensus=opening_prediction["metrics"]["consensus"],
            top_gap=opening_prediction["metrics"]["topGap"],
            phase_status=prediction["phaseStatus"],
            effective_mode=prediction["effectiveModeLabel"],
            final_action=prediction["finalAction"],
            final_prediction=prediction["finalPrediction"],
            decision_basis=prediction["decisionBasis"],
            mode_selection_basis=prediction["modeSelectionBasis"],
            mode_history_accuracy=prediction["modeHistoryAccuracy"],
            history_sample_size=prediction["historySampleSize"],
            history_accuracy_summary=prediction["historyAccuracySummary"],
            source_match_id=match.schedule_id,
            source_page_url=match.europe_odds_url,
            source_js_url=oddslist_js_url,
            opening_odds_json=json.dumps(
                {
                    item["company"]: item["openingOdds"]
                    for item in plugin_rows
                },
                ensure_ascii=False,
            ),
            closing_odds_json=json.dumps(
                {
                    item["company"]: item["closingOdds"]
                    for item in plugin_rows
                    if item.get("closingOdds")
                },
                ensure_ascii=False,
            ),
            explanation=opening_prediction["explanation"],
        ),
        audit_rows,
    )


def set_column_widths(sheet, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def style_header_row(sheet, row_index: int = 1) -> None:
    fill = workbook_header_fill()
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def append_match_rows(sheet, rows: Iterable[MatchReportRow]) -> None:
    for row in rows:
        sheet.append(
            [
                row.kickoff_text,
                row.home_team,
                row.away_team,
                row.recommendation,
                row.structure_label,
                row.confidence,
                round(row.consensus, 4),
                round(row.top_gap, 4),
                row.phase_status,
                row.effective_mode,
                row.final_action,
                row.final_prediction,
                row.decision_basis,
                row.mode_selection_basis,
                row.mode_history_accuracy,
                row.history_sample_size,
                row.history_accuracy_summary,
                row.source_match_id,
                row.source_page_url,
                row.source_js_url,
                row.opening_odds_json,
                row.closing_odds_json,
                row.explanation,
            ]
        )


def append_league_section(
    sheet,
    league: str,
    league_rows: list[MatchReportRow],
    detail_headers: list[str],
) -> None:
    section_row_index = sheet.max_row + 1
    sheet.append([league] + [""] * (len(detail_headers) - 1))
    for cell in sheet[section_row_index]:
        cell.font = Font(bold=True)
        cell.fill = section_fill()
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    header_row_index = sheet.max_row + 1
    sheet.append(detail_headers)
    style_header_row(sheet, header_row_index)
    append_match_rows(sheet, league_rows)


def append_betting_decision_table_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("投注决策总表")
    table_rows = get_betting_decision_table()
    headers = [
        "信任等级",
        "结构标签",
        "原始类型",
        "前二差值范围",
        "最终决策动作",
        "最终预测结果规则",
        "历史样本数",
        "历史准确率1",
        "历史准确率2",
        "历史准确率3",
        "规则说明",
        "样本说明",
    ]
    sheet.append(headers)
    style_header_row(sheet)
    for row in table_rows:
        sheet.append([row[header] for header in headers])
    set_column_widths(
        sheet,
        {
            1: 10,
            2: 16,
            3: 12,
            4: 14,
            5: 14,
            6: 24,
            7: 12,
            8: 22,
            9: 22,
            10: 22,
            11: 42,
            12: 34,
        },
    )
    sheet.freeze_panes = "A2"
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_live_mode_selection_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("临场模式选择表")
    table_rows = get_live_mode_selection_table()
    headers = [
        "结构标签",
        "前二差值范围",
        "基础决策类型",
        "推荐模式",
        "历史样本数",
        "初赔Top1",
        "初赔单选准确率",
        "初赔覆盖率",
        "临场Top1",
        "临场单选准确率",
        "临场覆盖率",
        "混合Top1",
        "混合单选准确率",
        "混合覆盖率",
        "推荐依据",
    ]
    sheet.append(headers)
    style_header_row(sheet)
    for row in table_rows:
        sheet.append([row[header] for header in headers])
    set_column_widths(
        sheet,
        {
            1: 16,
            2: 14,
            3: 12,
            4: 12,
            5: 12,
            6: 12,
            7: 16,
            8: 14,
            9: 12,
            10: 16,
            11: 14,
            12: 12,
            13: 16,
            14: 14,
            15: 42,
        },
    )
    sheet.freeze_panes = "A2"
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(rows: list[MatchReportRow], audit_rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"

    overview["A1"] = "筛选条件"
    overview["A1"].font = Font(bold=True)
    overview["A2"] = "时间窗口（北京时间）"
    overview["B2"] = f"{format_kickoff_bjt(WINDOW_START_BJT)} 至 {format_kickoff_bjt(WINDOW_END_BJT)}"
    overview["A3"] = "比赛范围"
    overview["B3"] = (
        f"球探未来赛程页 Next_{FUTURE_SCHEDULE_DATES[0]} 至 "
        f"Next_{FUTURE_SCHEDULE_DATES[-1]} 的足球比赛"
    )
    overview["A4"] = "赔率条件"
    overview["B4"] = "固定 6 家公司初赔齐全；若已到开球时间且临场赔率齐全，则在初赔/临场/初赔+临场三种模式中按保守约束自动择优。"
    overview["A5"] = "预测条件"
    overview["B5"] = f"输出预测信任等级为{format_confidence_scope()}的比赛"
    overview["A6"] = "运行时点（北京时间）"
    overview["B6"] = format_kickoff_bjt(CURRENT_TIME_BJT)

    phase_counts = {"未开赛": 0, "已开球": 0}
    mode_counts = {"初赔": 0, "临场": 0, "初赔+临场": 0}
    for row in rows:
        phase_counts[row.phase_status] = phase_counts.get(row.phase_status, 0) + 1
        mode_counts[row.effective_mode] = mode_counts.get(row.effective_mode, 0) + 1

    overview["A7"] = f"{format_confidence_scope()}信任比赛数"
    overview["B7"] = len(rows)
    overview["A8"] = "未开赛场次"
    overview["B8"] = phase_counts.get("未开赛", 0)
    overview["A9"] = "已开球场次"
    overview["B9"] = phase_counts.get("已开球", 0)
    overview["A10"] = "初赔模式生效场次"
    overview["B10"] = mode_counts.get("初赔", 0)
    overview["A11"] = "临场模式生效场次"
    overview["B11"] = mode_counts.get("临场", 0)
    overview["A12"] = "初赔+临场模式生效场次"
    overview["B12"] = mode_counts.get("初赔+临场", 0)
    overview["A13"] = "扫描备注数"
    overview["B13"] = len(audit_rows)

    overview["A15"] = "联赛"
    overview["B15"] = "场次"
    style_header_row(overview, 15)

    league_counts: dict[str, int] = {}
    for row in rows:
        league_counts[row.league] = league_counts.get(row.league, 0) + 1

    current_row = 16
    for league, count in sorted(league_counts.items()):
        overview.cell(current_row, 1, league)
        overview.cell(current_row, 2, count)
        current_row += 1

    set_column_widths(overview, {1: 22, 2: 24})

    grouped: dict[str, list[MatchReportRow]] = {}
    for row in rows:
        grouped.setdefault(row.league, []).append(row)

    detail_headers = [
        "开赛时间（北京时间）",
        "主队",
        "客队",
        "原始预测结果",
        "结构标签",
        "信任等级",
        "市场共识",
        "前二差值",
        "比赛时点状态",
        "生效预测模式",
        "最终决策动作",
        "最终预测结果",
        "决策依据",
        "模式选择依据",
        "模式历史准确率",
        "历史样本数",
        "历史准确率提示",
        "比赛ID",
        "欧赔页",
        "真实JS源",
        "六家公司初赔",
        "六家公司临场赔率",
        "说明",
    ]

    grouped_by_date: dict[str, list[MatchReportRow]] = {}
    for row in rows:
        grouped_by_date.setdefault(row.kickoff_bjt.strftime("%Y-%m-%d"), []).append(row)

    for date_text, date_rows in sorted(grouped_by_date.items()):
        sheet = workbook.create_sheet(normalize_sheet_name(date_text))
        league_groups: dict[str, list[MatchReportRow]] = {}
        for row in date_rows:
            league_groups.setdefault(row.league, []).append(row)

        sorted_league_items = sorted(
            league_groups.items(),
            key=lambda item: (
                min(row.kickoff_bjt for row in item[1]),
                item[0],
            ),
        )
        for league, league_rows in sorted_league_items:
            append_league_section(sheet, league, league_rows, detail_headers)
            sheet.append([""] * len(detail_headers))

        set_column_widths(
            sheet,
            {
                1: 20,
                2: 18,
                3: 18,
                4: 12,
                5: 14,
                6: 10,
                7: 10,
                8: 10,
                9: 12,
                10: 12,
                11: 14,
                12: 14,
                13: 28,
                14: 36,
                15: 36,
                16: 12,
                17: 36,
                18: 12,
                19: 40,
                20: 36,
                21: 60,
                22: 60,
                23: 42,
            },
        )
        sheet.freeze_panes = "A1"
        for row_cells in sheet.iter_rows(min_row=1):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    append_betting_decision_table_sheet(workbook)
    append_live_mode_selection_sheet(workbook)

    audit_sheet = workbook.create_sheet("扫描备注")
    audit_sheet.append(["比赛ID", "状态", "详情"])
    style_header_row(audit_sheet)
    if audit_rows:
        for item in audit_rows:
            audit_sheet.append([item["schedule_id"], item["status"], item["detail"]])
    set_column_widths(audit_sheet, {1: 12, 2: 24, 3: 80})
    audit_sheet.freeze_panes = "A2"
    for row_cells in audit_sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if not rows:
        notice_sheet = workbook.create_sheet("结果说明")
        notice_sheet["A1"] = (
            f"本次按当前条件筛选后，没有命中{format_confidence_scope()}信任比赛。"
        )
        notice_sheet["A2"] = "你仍可在“扫描备注”Sheet 查看每场被排除的原因。"
        notice_sheet["A1"].font = Font(bold=True)
        notice_sheet["A1"].fill = section_fill()
        set_column_widths(notice_sheet, {1: 56})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    global WINDOW_START_BJT, WINDOW_END_BJT, FUTURE_SCHEDULE_DATES, ALLOWED_CONFIDENCES, CURRENT_TIME_BJT

    parser = argparse.ArgumentParser(
        description="根据北京时间范围抓取球探未来赛程，生成开球感知的欧赔预测 Excel。"
    )
    parser.add_argument(
        "--start",
        default="2026-03-29 18:00",
        help="开始时间，格式：YYYY-MM-DD HH:MM，按北京时间解释。",
    )
    parser.add_argument(
        "--end",
        default="2026-04-03 00:00",
        help="结束时间，格式：YYYY-MM-DD HH:MM，按北京时间解释。",
    )
    parser.add_argument(
        "--confidences",
        default="高,中",
        help="输出的信任等级，逗号分隔，例如：高、高,中、高,中,谨慎",
    )
    parser.add_argument(
        "--now",
        help="运行时点，格式：YYYY-MM-DD HH:MM，按北京时间解释；不传则使用当前北京时间。",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="输出 Excel 路径；不传则按时间范围自动命名。",
    )
    args = parser.parse_args()

    WINDOW_START_BJT = parse_datetime_bjt(args.start)
    WINDOW_END_BJT = parse_datetime_bjt(args.end)
    if WINDOW_END_BJT < WINDOW_START_BJT:
        raise ValueError("结束时间不能早于开始时间。")
    FUTURE_SCHEDULE_DATES = build_future_schedule_dates(WINDOW_START_BJT, WINDOW_END_BJT)
    CURRENT_TIME_BJT = (
        parse_datetime_bjt(args.now)
        if args.now
        else datetime.now(TIMEZONE_BEIJING)
    )
    ALLOWED_CONFIDENCES = {
        item.strip()
        for item in args.confidences.split(",")
        if item.strip()
    }
    invalid_confidences = sorted(ALLOWED_CONFIDENCES - VALID_CONFIDENCES)
    if invalid_confidences:
        raise ValueError(
            f"不支持的信任等级: {', '.join(invalid_confidences)}；仅支持 {', '.join(sorted(VALID_CONFIDENCES))}。"
        )

    rows, audit_rows = iter_report_rows()
    if args.output:
        output_path = args.output
    else:
        confidence_slug = "_".join(sorted(ALLOWED_CONFIDENCES))
        output_path = OUTPUT_DIR / (
            f"titan007_{confidence_slug}_confidence_"
            f"{WINDOW_START_BJT.strftime('%Y%m%d_%H%M')}_"
            f"{WINDOW_END_BJT.strftime('%Y%m%d_%H%M')}.xlsx"
        )
    write_workbook(rows, audit_rows, output_path)
    print(output_path)
    print(json.dumps({"high_medium_confidence_matches": len(rows), "audit_rows": len(audit_rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
